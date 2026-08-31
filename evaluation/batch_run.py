# -*- coding: utf-8 -*-
"""批量运行 single_test 逻辑并回填 Excel"""
import sys, os, time, math

# 确保能导入 evaluate_localization
sys.path.insert(0, r'D:\TTT\GeoKG\Helsinki data\test1\Karlsruhe_test')

from evaluate_localization import BatchLocalizer
import pandas as pd
from openpyxl import load_workbook

EXCEL_PATH = r'D:\TTT\GeoKG\Helsinki data\test1\Karlsruhe_test\KG实验结果统计.xlsx'
SHEET_NAME = 'delaunay'

# Haversine 距离（米）
def haversine(lon1, lat1, lon2, lat2):
    R = 6371000
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlambda = math.radians(lon2 - lon1)
    a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlambda/2)**2
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
    return R * c

def parse_coord(s):
    """解析 'lat，lon' 或 'lat,lon' 字符串"""
    if pd.isna(s):
        return None, None
    s = str(s).strip()
    for sep in ['，', ',']:
        if sep in s:
            parts = s.split(sep)
            if len(parts) >= 2:
                try:
                    lat = float(parts[0].strip())
                    lon = float(parts[1].strip())
                    return lon, lat
                except ValueError:
                    continue
    return None, None

def main():
    df = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_NAME)
    print(f"读取到 {len(df)} 行数据")

    localizer = BatchLocalizer()

    results = []
    for idx, row in df.iterrows():
        desc = row.get('描述', '')
        if pd.isna(desc) or str(desc).strip() == '':
            print(f"[{idx+1}/30] 跳过空描述")
            results.append({})
            continue

        desc = str(desc).strip()
        print(f"\n[{idx+1}/30] 正在定位...")
        print(f"描述: {desc[:80]}...")

        t0 = time.time()
        result = localizer.localize_single(desc)
        elapsed = time.time() - t0

        pred_lon, pred_lat = None, None
        confidence = 0.0
        if result['status'] == 'success':
            pred_lon = result.get('lon')
            pred_lat = result.get('lat')
            confidence = result.get('confidence', 0)
            print(f"  -> 成功: lon={pred_lon:.6f}, lat={pred_lat:.6f}, conf={confidence:.2%}, time={elapsed:.1f}s")
        else:
            print(f"  -> 失败: {result.get('message', '')}, time={elapsed:.1f}s")

        # 计算误差
        real_coord = row.get('真实坐标', '')
        real_lon, real_lat = parse_coord(real_coord)
        distance = None
        if real_lon is not None and real_lat is not None and pred_lon is not None and pred_lat is not None:
            distance = haversine(real_lon, real_lat, pred_lon, pred_lat)
            print(f"  -> 直线距离: {distance:.1f}m")

        results.append({
            'pred_coord': f"{pred_lon:.6f}, {pred_lat:.6f}" if pred_lon is not None and pred_lat is not None else None,
            'distance': round(distance, 2) if distance is not None else None,
            'confidence': round(confidence, 4) if confidence else None,
            'time': round(elapsed, 2),
        })

    localizer.close()

    # 写回 Excel（保留原有格式）
    print("\n正在写回 Excel...")
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]

    for idx, res in enumerate(results):
        excel_row = idx + 2  # 数据从第2行开始
        if res.get('pred_coord') is not None:
            ws.cell(row=excel_row, column=3, value=res['pred_coord'])  # C列=预测坐标
        if res.get('distance') is not None:
            ws.cell(row=excel_row, column=4, value=res['distance'])    # D列=误差
        if res.get('confidence') is not None:
            ws.cell(row=excel_row, column=5, value=res['confidence'])  # E列=置信度
        if res.get('time') is not None:
            ws.cell(row=excel_row, column=7, value=res['time'])        # G列=时间

    wb.save(EXCEL_PATH)
    print(f"已保存: {EXCEL_PATH}")

if __name__ == '__main__':
    main()
