# -*- coding: utf-8 -*-
"""
No-KG Attribute + Geometry baseline for Karlsruhe localization ablation.

This script does not modify existing source code. It reads 28 samples from
KG实验结果统计.xlsx / 50m and uses the already generated scene descriptions in
scenes_100_llm.json as the structured LLM-description source. It does NOT use
KG edge files, Neo4j, building-building relations, or building-POI relations.

Outputs are written back to KG实验结果统计.xlsx:
- no_kg
- kg_vs_no_kg_metrics
"""

from __future__ import annotations

import json
import math
import time
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from scipy.spatial import cKDTree
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

ROOT = Path(r"D:\TTT\GeoKG\Helsinki data\test1")
TEST_DIR = ROOT / "Karlsruhe_test"
DATA_DIR = ROOT / "Karlsruhe-data-prepare"
KGDATA_DIR = DATA_DIR / "KGdata" / "r100"
EXCEL_PATH = TEST_DIR / "KG实验结果统计.xlsx"
SCENES_JSON = DATA_DIR / "scenes_100_llm.json"
BUILDINGS_PATH = KGDATA_DIR / "buildings.csv"
POIS_PATH = KGDATA_DIR / "pois.csv"

INPUT_SHEET = "50m"
KG_SHEET = "100m"
OUTPUT_SHEET = "no_kg"
METRIC_SHEET = "kg_vs_no_kg_metrics"
N_SAMPLES = 28
POI_SEARCH_RADIUS_M = 35.0
MAX_CANDIDATES_PER_ENTITY = 80
MAX_COMBINATIONS = 8000
EARTH_RADIUS = 6371000
DIRECTIONS_8 = ["E", "NE", "N", "NW", "W", "SW", "S", "SE"]
DIRECTION_ANGLE = {"E": 0, "NE": 45, "N": 90, "NW": 135, "W": 180, "SW": 225, "S": 270, "SE": 315}
OPPOSITE_DIRECTION = {"N": "S", "S": "N", "E": "W", "W": "E", "NE": "SW", "SW": "NE", "NW": "SE", "SE": "NW"}


@dataclass
class RefEntity:
    entity_id: str
    color_side: str
    color_top: str
    direction_from_user: str
    estimated_distance: float
    poi_types: list[str]


@dataclass
class NoKGCandidate:
    query_id: str
    building_id: str
    lon: float
    lat: float
    color_side: str
    color_top: str
    score: float
    poi_satisfied: int
    poi_total: int
    matched_poi_types: list[str]


def haversine_distance(lat1, lon1, lat2, lon2):
    lat1, lon1, lat2, lon2 = map(math.radians, [lat1, lon1, lat2, lon2])
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = math.sin(dlat / 2) ** 2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon / 2) ** 2
    return EARTH_RADIUS * 2 * math.asin(math.sqrt(a))


def direction_to_offset(direction, distance_meters, lat=49.0):
    angle = math.radians(DIRECTION_ANGLE.get(direction, 0))
    meters_per_deg_lon = 111000 * math.cos(math.radians(lat))
    return (distance_meters * math.cos(angle)) / meters_per_deg_lon, (distance_meters * math.sin(angle)) / 111000


def compute_angle_deg(lon1, lat1, lon2, lat2):
    return (math.degrees(math.atan2(lat2 - lat1, lon2 - lon1)) + 360) % 360


def compute_direction_8(angle_deg):
    return DIRECTIONS_8[int(((angle_deg + 22.5) % 360) / 45)]


def calculate_possible_directions(dir1, dist1, dir2, dist2):
    if not dir1 or not dir2 or dist1 <= 0 or dist2 <= 0:
        return []
    x1, y1 = dist1 * math.cos(math.radians(DIRECTION_ANGLE.get(dir1, 0))), dist1 * math.sin(math.radians(DIRECTION_ANGLE.get(dir1, 0)))
    x2, y2 = dist2 * math.cos(math.radians(DIRECTION_ANGLE.get(dir2, 0))), dist2 * math.sin(math.radians(DIRECTION_ANGLE.get(dir2, 0)))
    angle = (math.degrees(math.atan2(y2 - y1, x2 - x1)) + 360) % 360
    main_idx = int(((angle + 22.5) % 360) / 45)
    dist_ratio = max(dist1, dist2) / min(dist1, dist2) if min(dist1, dist2) > 0 else 1
    tolerance = 1 if dist_ratio < 1.5 else 2 if dist_ratio < 3 else 3
    dirs = []
    for i in range(-tolerance, tolerance + 1):
        dirs.append(DIRECTIONS_8[(main_idx + i) % 8])
    main_dir = DIRECTIONS_8[main_idx]
    if main_dir in dirs:
        dirs.remove(main_dir)
    return [main_dir] + dirs


def parse_coord(text: Any):
    if pd.isna(text):
        return None, None
    parts = [p.strip() for p in str(text).replace("，", ",").split(",")]
    if len(parts) < 2:
        return None, None
    a, b = float(parts[0]), float(parts[1])
    if 45 <= a <= 55 and 5 <= b <= 15:
        return b, a
    return a, b


def load_samples():
    df = pd.read_excel(EXCEL_PATH, sheet_name=INPUT_SHEET).iloc[:N_SAMPLES].copy()
    rows = []
    for _, r in df.iterrows():
        lon, lat = parse_coord(r["真实坐标"])
        rows.append({"点id": int(r["点id"]), "真实坐标": r["真实坐标"], "true_lon": lon, "true_lat": lat, "描述": r["描述"]})
    return pd.DataFrame(rows)


def parse_distance(text):
    words = {
        "one": 1, "two": 2, "three": 3, "four": 4, "five": 5, "six": 6, "seven": 7, "eight": 8, "nine": 9,
        "ten": 10, "eleven": 11, "twelve": 12, "thirteen": 13, "fourteen": 14, "fifteen": 15, "sixteen": 16,
        "seventeen": 17, "eighteen": 18, "nineteen": 19, "twenty": 20, "twenty-five": 25, "thirty": 30,
        "thirty-five": 35, "forty": 40, "fifty": 50,
    }
    import re
    m = re.search(r"(\d+(?:\.\d+)?)\s*meters?", text, re.I)
    if m:
        return float(m.group(1))
    for k, v in words.items():
        if re.search(rf"\b{k}\b\s+meters?", text, re.I):
            return float(v)
    return 20.0


def normalize_direction(text):
    import re
    patterns = [
        ("northeast|north-east", "NE"), ("northwest|north-west", "NW"),
        ("southeast|south-east", "SE"), ("southwest|south-west", "SW"),
        ("north", "N"), ("south", "S"), ("east", "E"), ("west", "W"),
    ]
    for pat, val in patterns:
        if re.search(pat, text, re.I):
            return val
    return ""


def parse_colors(text):
    import re
    side, top = "unknown", "unknown"
    m = re.search(r"with\s+(.+?)\s+sides?\s+and\s+(.+?)\s+roof", text, re.I)
    if not m:
        m = re.search(r"with\s+a\s+(.+?)\s+side\s+and\s+(.+?)\s+roof", text, re.I)
    if m:
        side, top = m.group(1).strip(), m.group(2).strip()
    return side.lower(), top.lower()


def parse_pois(text):
    import re
    tokens = []
    for key in ["housing", "containing", "contains", "has", "inside which there is", "where"]:
        for m in re.finditer(key + r"\s+([^.;]+)", text, re.I):
            frag = m.group(1).lower()
            frag = re.sub(r"\b(a|an|the|two|three|and|alongside|inside|which|there|is|are|with)\b", " ", frag)
            for t in re.split(r"[,/]|\s+and\s+", frag):
                t = t.strip().replace(" ", "_")
                if t and len(t) > 2:
                    tokens.append(t)
    return sorted(set(tokens))


def parse_description_refs(description):
    import re
    parts = re.split(r"(?<=[.;])\s+|;\s+", str(description))
    refs = []
    for part in parts:
        if "building" not in part.lower():
            continue
        direction = normalize_direction(part)
        if not direction:
            continue
        dist = parse_distance(part)
        side, top = parse_colors(part)
        pois = parse_pois(part)
        refs.append(RefEntity(f"ref_{len(refs)}", side, top, direction, dist, pois))
    if not refs and "building" in str(description).lower():
        direction = normalize_direction(description)
        side, top = parse_colors(description)
        refs.append(RefEntity("ref_0", side, top, direction or "N", parse_distance(description), parse_pois(description)))
    return refs


def load_data():
    buildings = pd.read_csv(BUILDINGS_PATH, encoding="utf-8-sig").rename(columns={"id:ID": "id"})
    pois = pd.read_csv(POIS_PATH, encoding="utf-8-sig").rename(columns={"id:ID": "id"})
    buildings = buildings.dropna(subset=["id", "lon", "lat"]).copy()
    pois = pois.dropna(subset=["id", "lon", "lat", "fclass"]).copy()
    buildings["lon"] = buildings["lon"].astype(float)
    buildings["lat"] = buildings["lat"].astype(float)
    buildings = buildings[buildings["lon"].between(8.34, 8.45) & buildings["lat"].between(48.98, 49.05)].copy()
    pois["lon"] = pois["lon"].astype(float)
    pois["lat"] = pois["lat"].astype(float)
    return buildings, pois


def color_match_score(value: str, target: str):
    value, target = str(value or "").lower().strip(), str(target or "").lower().strip()
    if not value or value == "unknown":
        return 0.20
    if not target or target == "unknown":
        return 0.08
    if value == target:
        return 1.0
    if value in target or target in value:
        return 0.70
    if set(value.replace("-", " ").split()) & set(target.replace("-", " ").split()):
        return 0.40
    return 0.0


def build_poi_index(pois):
    index = {}
    for fclass, group in pois.groupby(pois["fclass"].astype(str).str.lower()):
        coords = group[["lon", "lat"]].to_numpy(float)
        if len(coords):
            index[fclass] = cKDTree(coords)
    return index


def count_nearby_pois(building, poi_index, poi_types):
    if not poi_types:
        return 0, []
    matched = []
    lon, lat = float(building["lon"]), float(building["lat"])
    deg_radius = POI_SEARCH_RADIUS_M / 111000.0
    for poi_type in poi_types:
        tree = poi_index.get(str(poi_type).lower())
        if tree is not None and tree.query_ball_point([lon, lat], deg_radius):
            matched.append(poi_type)
    return len(matched), matched


def search_entity_candidates(entity, buildings, poi_index):
    df = buildings.copy()
    if entity.color_side and entity.color_side.lower() != "unknown":
        side = entity.color_side.lower()
        df = df[df["color_side"].astype(str).str.lower().str.contains(side, na=False) |
                df["color_top"].astype(str).str.lower().str.contains(side, na=False)]
    elif entity.color_top and entity.color_top.lower() != "unknown":
        top = entity.color_top.lower()
        df = df[df["color_top"].astype(str).str.lower().str.contains(top, na=False) |
                df["color_side"].astype(str).str.lower().str.contains(top, na=False)]
    if df.empty:
        df = buildings
    rows = []
    # Random-looking but deterministic spatial thinning: preserve broad candidates without full scan explosion.
    if len(df) > 4000:
        df = df.iloc[::max(1, len(df) // 4000)].copy()
    for b in df.itertuples(index=False):
        bd = b._asdict()
        side_score = color_match_score(entity.color_side, bd.get("color_side", ""))
        top_score = color_match_score(entity.color_top, bd.get("color_top", ""))
        if side_score <= 0 and top_score <= 0:
            continue
        poi_sat, matched_pois = count_nearby_pois(bd, poi_index, entity.poi_types)
        if entity.poi_types and poi_sat == 0:
            continue
        score = side_score + top_score + poi_sat * 2.5 + (0.2 if entity.estimated_distance > 0 else 0)
        rows.append(NoKGCandidate(entity.entity_id, str(bd["id"]), float(bd["lon"]), float(bd["lat"]), str(bd.get("color_side", "")), str(bd.get("color_top", "")), float(score), poi_sat, len(entity.poi_types), matched_pois))
    rows.sort(key=lambda c: (-c.poi_satisfied, -c.score))
    return rows[:MAX_CANDIDATES_PER_ENTITY]


def estimate_position(candidates, entities):
    cmap = {c.query_id: c for c in candidates}
    positions, weights = [], []
    for e in entities:
        c = cmap.get(e.entity_id)
        if not c or not e.direction_from_user:
            continue
        dist = e.estimated_distance if e.estimated_distance > 0 else 20.0
        opposite = OPPOSITE_DIRECTION.get(e.direction_from_user, e.direction_from_user)
        dlon, dlat = direction_to_offset(opposite, dist, c.lat)
        positions.append((c.lon + dlon, c.lat + dlat))
        weights.append(1.0 / max(dist, 5.0))
    if not positions:
        return None, None, 0.0
    total = sum(weights)
    lon = sum(p[0] * w for p, w in zip(positions, weights)) / total
    lat = sum(p[1] * w for p, w in zip(positions, weights)) / total
    spread = np.mean([haversine_distance(lat, lon, p[1], p[0]) for p in positions]) if len(positions) > 1 else 60.0
    geom_conf = max(0.0, 1.0 - spread / 120.0)
    return lon, lat, geom_conf


def relation_consistency(cands, entities):
    cmap = {c.query_id: c for c in cands}
    vals = []
    for i, e1 in enumerate(entities):
        for e2 in entities[i + 1:]:
            c1, c2 = cmap.get(e1.entity_id), cmap.get(e2.entity_id)
            if not c1 or not c2:
                continue
            dirs = calculate_possible_directions(e1.direction_from_user, e1.estimated_distance, e2.direction_from_user, e2.estimated_distance)
            if not dirs:
                continue
            actual = compute_direction_8(compute_angle_deg(c1.lon, c1.lat, c2.lon, c2.lat))
            vals.append(1.0 if actual in dirs else 0.0)
    return float(np.mean(vals)) if vals else 0.5


def build_combinations(candidate_lists, entities):
    combos = [([], set(), 0.0)]
    for cand_list in candidate_lists:
        new = []
        for existing, used, score in combos:
            for cand in cand_list:
                if cand.building_id in used:
                    continue
                new.append((existing + [cand], used | {cand.building_id}, score + cand.score))
        new.sort(key=lambda x: -x[2])
        combos = new[:MAX_COMBINATIONS]
    results = []
    for cands, _, score in combos:
        lon, lat, geom_conf = estimate_position(cands, entities)
        if lon is None:
            continue
        rel_score = relation_consistency(cands, entities)
        poi_total = sum(c.poi_total for c in cands)
        poi_sat = sum(c.poi_satisfied for c in cands)
        poi_score = poi_sat / poi_total if poi_total else 0.8
        total = score + rel_score * 1.2 + poi_score * 1.5 + geom_conf * 2.0
        confidence = max(0.0, min(1.0, 0.45 * geom_conf + 0.25 * poi_score + 0.20 * rel_score + 0.10 * min(score / max(len(cands) * 5, 1), 1)))
        results.append((total, confidence, lon, lat, cands))
    results.sort(key=lambda x: (-x[0], -x[1]))
    return results[:50]


def localize_one(description, buildings, poi_index):
    entities = parse_description_refs(description)
    if not entities:
        return {"status": "no_reference", "message": "无结构化参照物"}
    lists = []
    for e in entities:
        cands = search_entity_candidates(e, buildings, poi_index)
        if not cands:
            return {"status": "no_match", "message": f"未找到候选: {e.entity_id}", "num_references": len(entities)}
        lists.append(cands)
    combos = build_combinations(lists, entities)
    if not combos:
        return {"status": "no_position", "message": "无法形成候选组合", "num_references": len(entities)}
    score, conf, lon, lat, cands = combos[0]
    matched = [{"desc_id": c.query_id, "building_id": c.building_id, "colors": f"{c.color_side}/{c.color_top}", "matched_pois": ",".join(c.matched_poi_types)} for c in cands]
    return {"status": "success", "lon": lon, "lat": lat, "confidence": conf, "num_references": len(entities), "num_matches": len(combos), "matched_combo": matched}


def summarize_metrics(df, confidence_col="置信度"):
    n = len(df)
    work = df.copy()
    work["误差"] = pd.to_numeric(work["误差"], errors="coerce")
    if confidence_col in work:
        work[confidence_col] = pd.to_numeric(work[confidence_col], errors="coerce")
    success = work[work["误差"].notna()].copy()
    errors = success["误差"].astype(float).to_numpy()
    confs = success[confidence_col].dropna().astype(float).to_numpy() if confidence_col in success else np.array([])
    return {
        "Acc@5m": float((errors <= 5).sum() / n) if n else 0,
        "Acc@10m": float((errors <= 10).sum() / n) if n else 0,
        "Acc@15m": float((errors <= 15).sum() / n) if n else 0,
        "Mean error": float(np.mean(errors)) if len(errors) else None,
        "Median error": float(np.median(errors)) if len(errors) else None,
        "Mean confidence": float(np.mean(confs)) if len(confs) else None,
        "Success count": int(len(success)),
        "Total count": int(n),
    }


def build_metric_sheet(no_kg_df):
    kg_df = pd.read_excel(EXCEL_PATH, sheet_name=KG_SHEET).iloc[:N_SAMPLES].copy()
    kg_metric = summarize_metrics(kg_df)
    no_metric = summarize_metrics(no_kg_df)
    rows = []
    for method, used_kg, graph_rel, metric in [
        ("GeoKG-Loc (Buffer radius=100m)", "Yes", "Yes", kg_metric),
        ("GeoLoc w/o KG (Attribute + Geometry)", "No", "No", no_metric),
    ]:
        rows.append({
            "Method": method,
            "KG used": used_kg,
            "Graph relation used": graph_rel,
            "Acc@5m": metric["Acc@5m"],
            "Acc@10m": metric["Acc@10m"],
            "Acc@15m": metric["Acc@15m"],
            "Mean error (m)": metric["Mean error"],
            "Median error (m)": metric["Median error"],
            "Mean confidence": metric["Mean confidence"],
            "Success/Total": f"{metric['Success count']}/{metric['Total count']}",
        })
    return pd.DataFrame(rows)


def write_excel(no_kg_df, metrics_df):
    target_path = EXCEL_PATH
    try:
        with pd.ExcelWriter(target_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            no_kg_df.to_excel(writer, sheet_name=OUTPUT_SHEET, index=False)
            metrics_df.to_excel(writer, sheet_name=METRIC_SHEET, index=False)
    except PermissionError:
        target_path = TEST_DIR / "KG实验结果统计_no_kg.xlsx"
        shutil.copy2(EXCEL_PATH, target_path)
        with pd.ExcelWriter(target_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            no_kg_df.to_excel(writer, sheet_name=OUTPUT_SHEET, index=False)
            metrics_df.to_excel(writer, sheet_name=METRIC_SHEET, index=False)
    wb = load_workbook(target_path)
    fill = PatternFill("solid", fgColor="EAF2FF")
    font = Font(bold=True, color="1F2937")
    thin = Side(style="thin", color="D9DEE7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ws_name in [OUTPUT_SHEET, METRIC_SHEET]:
        ws = wb[ws_name]
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 60)
    wb.save(target_path)
    return target_path


def main():
    samples = load_samples()
    buildings, pois = load_data()
    poi_index = build_poi_index(pois)
    rows = []
    for _, s in samples.iterrows():
        t0 = time.time()
        r = localize_one(s["描述"], buildings, poi_index)
        elapsed = time.time() - t0
        pred, err, conf, matched = "", None, None, ""
        if r["status"] == "success":
            pred = f"{r['lon']:.6f}, {r['lat']:.6f}"
            err = round(haversine_distance(s["true_lat"], s["true_lon"], r["lat"], r["lon"]), 2)
            conf = round(float(r["confidence"]), 4)
            matched = json.dumps(r["matched_combo"], ensure_ascii=False)
        rows.append({
            "点id": s["点id"], "真实坐标": s["真实坐标"], "预测坐标": pred,
            "误差": err, "置信度": conf, "描述": s["描述"], "时间": round(elapsed, 2),
            "状态": r["status"], "参照物数量": r.get("num_references"), "候选组合数": r.get("num_matches"),
            "匹配详情": matched, "备注": r.get("message", ""),
        })
        print(f"point {s['点id']}: {r['status']} error={err} time={elapsed:.1f}s")
    no_kg_df = pd.DataFrame(rows)
    metrics_df = build_metric_sheet(no_kg_df)
    written_path = write_excel(no_kg_df, metrics_df)
    print(f"Written sheets: {OUTPUT_SHEET}, {METRIC_SHEET}")
    print(f"Workbook: {written_path}")
    print(metrics_df.to_string(index=False))


if __name__ == "__main__":
    main()
